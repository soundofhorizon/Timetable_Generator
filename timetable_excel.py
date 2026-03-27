from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, cast

from timetable_tool import (
    CLASS_ORDER_DEFAULT,
    DAY_JP,
    DAY_ORDER,
    ONBI_FILL_COLOR,
    TECHKA_FILL_COLOR,
    Slot,
    _class_sort_key,
    _normalize_subject_name,
    _origin_group_key,
    grade_of,
    parse_slot_key,
)


def to_excel_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def slot_to_col(slot: Slot, day_periods: Dict[str, int]) -> int:
    col = _slot_to_column(slot.day, slot.period, day_periods)
    if col is None:
        raise ValueError(f"invalid slot: {slot}")
    return col


def row_base_for_block(block: str) -> int:
    if block == "upper":
        return 3
    if block == "lower":
        return 17
    raise ValueError("block must be 'upper' or 'lower'")


SLOT_AREA_START_COL = 4
SLOT_AREA_END_COL = 32
RIGHT_LABEL_COL = 33
RIGHT_AUX_COL = 34
PREFERRED_GAP_AFTER = ("Mon", "Wed", "Tue", "Thu")
BLOCK_HEIGHT = 50
BLOCK_PITCH = BLOCK_HEIGHT + 1
TEACHER_ROW_START = 4
TEACHER_ROW_END = 31
CLASS_ROW_START = 35
CLASS_ROW_END = 47
SEPARATOR_ROW = 48
SPECIAL_ROW = 49
PATROL_ROW = 50
BOTTOM_ROW = 50
STATIC_BLOCK_MERGES = [
    "A1:C1",
    "D1:AH1",
    "A2:C3",
    "AG2:AH3",
    "A33:C34",
    "AG33:AH34",
    "A35:B35",
    "A36:B36",
    "A37:B37",
    "A38:B38",
    "A39:B39",
    "A40:B40",
    "A41:B41",
    "A42:B42",
    "A43:B43",
    "A44:B44",
    "A45:B45",
    "A46:B46",
    "A47:B47",
    "A49:C49",
    "AG49:AH49",
    "A50:C50",
    "AG50:AH50",
]
DAY_LETTERS = {"Mon": "A", "Tue": "B", "Wed": "C", "Thu": "D", "Fri": "E"}
EXCEL_UNAVAIL_HARD_MARK = "✕"
EXCEL_UNAVAIL_SOFT_MARK = "△"
EXCEL_UNAVAIL_HARD_FILL_COLOR = "FDE2E2"
EXCEL_UNAVAIL_SOFT_FILL_COLOR = "FAFAD2"


def _normalize_day_periods(day_periods: Dict[str, int]) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for day in DAY_ORDER:
        try:
            normalized[day] = max(0, int(day_periods.get(day, 0)))
        except Exception:
            normalized[day] = 0
    return normalized


def _build_sheet_layout(day_periods: Dict[str, int]) -> dict:
    counts = _normalize_day_periods(day_periods)
    capacity = SLOT_AREA_END_COL - SLOT_AREA_START_COL + 1
    total_slots = sum(counts.values())
    if total_slots > capacity:
        raise ValueError(
            f"授業コマ数が多すぎます: {total_slots}コマ（最大 {capacity} コマ）"
        )

    gap_budget = min(capacity - total_slots, max(0, len(DAY_ORDER) - 1))
    gap_after = {day: 0 for day in DAY_ORDER}
    for day in PREFERRED_GAP_AFTER:
        if gap_budget <= 0:
            break
        if counts.get(day, 0) > 0:
            gap_after[day] = 1
            gap_budget -= 1

    slot_columns: list[tuple[str, int, int]] = []
    day_ranges: dict[str, tuple[int, int] | None] = {day: None for day in DAY_ORDER}
    gap_columns: list[int] = []
    current_col = SLOT_AREA_START_COL

    for idx, day in enumerate(DAY_ORDER):
        day_count = counts[day]
        if day_count > 0:
            start_col = current_col
            end_col = start_col + day_count - 1
            day_ranges[day] = (start_col, end_col)
            for period in range(1, day_count + 1):
                slot_columns.append((day, period, start_col + period - 1))
            current_col = end_col + 1

        if idx < len(DAY_ORDER) - 1 and gap_after.get(day, 0):
            for _ in range(gap_after[day]):
                gap_columns.append(current_col)
                current_col += 1

    return {
        "counts": counts,
        "total_slots": total_slots,
        "slot_columns": slot_columns,
        "day_ranges": day_ranges,
        "gap_columns": gap_columns,
        "slot_area_start_col": SLOT_AREA_START_COL,
        "slot_area_end_col": SLOT_AREA_END_COL,
        "right_label_col": RIGHT_LABEL_COL,
        "right_aux_col": RIGHT_AUX_COL,
    }


def _iter_block_merges(day_periods: Dict[str, int]) -> list[str]:
    merges = list(STATIC_BLOCK_MERGES)
    layout = _build_sheet_layout(day_periods)
    for row in (2, 33):
        for day in DAY_ORDER:
            day_range = layout["day_ranges"].get(day)
            if not day_range:
                continue
            start_col, end_col = day_range
            merges.append(f"{to_excel_col(start_col)}{row}:{to_excel_col(end_col)}{row}")
    return merges


def _slot_to_column(day: str, period: int, day_periods: Dict[str, int]) -> int | None:
    layout = _build_sheet_layout(day_periods)
    day_range = layout["day_ranges"].get(day)
    if not day_range:
        return None
    start_col, end_col = day_range
    if 1 <= period <= (end_col - start_col + 1):
        return start_col + period - 1
    return None


def _iter_slot_columns(day_periods: Dict[str, int]) -> list[tuple[str, int, int]]:
    return list(_build_sheet_layout(day_periods)["slot_columns"])


def _apply_dimensions(ws, slot_width: float, col_c_width: float, day_periods: Dict[str, int]) -> None:
    from openpyxl.utils import get_column_letter

    layout = _build_sheet_layout(day_periods)
    gap_letters = {get_column_letter(col) for col in layout["gap_columns"]}
    widths = {
        "A": 5.5,
        "B": 6.0,
        "C": col_c_width,
        "AG": 6.0,
        "AH": 6.0,
    }
    for col in range(1, 35):
        letter = get_column_letter(col)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
        elif letter in gap_letters:
            ws.column_dimensions[letter].width = 1.0
        else:
            ws.column_dimensions[letter].width = slot_width


def _apply_row_heights(ws, start_row: int, grid_row_height: float) -> None:
    fixed = {
        1: 15.0,
        2: 12.75,
        3: 12.75,
        32: 0.75,
        33: 12.0,
        34: 12.0,
        SEPARATOR_ROW: 4.5,
    }
    for rel_row, height in fixed.items():
        ws.row_dimensions[start_row + (rel_row - 1)].height = height
    for rel_row in list(range(TEACHER_ROW_START, TEACHER_ROW_END + 1)) + list(
        range(CLASS_ROW_START, CLASS_ROW_END + 1)
    ) + [SPECIAL_ROW, PATROL_ROW, BOTTOM_ROW]:
        ws.row_dimensions[start_row + (rel_row - 1)].height = grid_row_height


def _merge_block(ws, start_row: int, day_periods: Dict[str, int]) -> None:
    from openpyxl.utils.cell import coordinate_from_string

    for rng in _iter_block_merges(day_periods):
        start, end = rng.split(":")
        s_col, s_row = coordinate_from_string(start)
        e_col, e_row = coordinate_from_string(end)
        s_row = int(s_row) + (start_row - 1)
        e_row = int(e_row) + (start_row - 1)
        ws.merge_cells(f"{s_col}{s_row}:{e_col}{e_row}")


def _apply_merge_borders(ws, start_row: int, border, day_periods: Dict[str, int]) -> None:
    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.utils import column_index_from_string

    for rng in _iter_block_merges(day_periods):
        start, end = rng.split(":")
        s_col, s_row = coordinate_from_string(start)
        e_col, e_row = coordinate_from_string(end)
        s_row = int(s_row) + (start_row - 1)
        e_row = int(e_row) + (start_row - 1)
        s_col_idx = column_index_from_string(s_col)
        e_col_idx = column_index_from_string(e_col)
        for r in range(s_row, e_row + 1):
            for c in range(s_col_idx, e_col_idx + 1):
                cell = ws.cell(r, c)
                cell.border = border


def _safe_cell(ws, row: int, col: int):
    cell = ws.cell(row, col)
    if cell.__class__.__name__ == "MergedCell":
        return None
    return cell


def _set_value(cell, value) -> None:
    if cell is None:
        return
    if cell.__class__.__name__ == "MergedCell":
        return
    cell.value = value


def _init_block(
    ws,
    start_row: int,
    scenario_label: str,
    title_text: str,
    day_periods: Dict[str, int],
    use_prefix: bool = True,
    grid_row_height: float = 15.0,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side

    layout = _build_sheet_layout(day_periods)
    _merge_block(ws, start_row, day_periods)
    _apply_row_heights(ws, start_row, grid_row_height)

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_bottom = Border(bottom=thin)
    border_left_bottom = Border(left=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")

    font_date = Font(name="Arial", size=14)
    font_title = Font(name="HG正楷書体-PRO", size=12)
    font_scenario = Font(name="HG創英角ﾎﾟｯﾌﾟ体", size=20)
    font_day = Font(name="ＭＳ 明朝", size=12)
    font_time = Font(name="HG丸ｺﾞｼｯｸM-PRO", size=12)
    font_grid = Font(name="游ゴシック", size=10)
    font_grid_alt = Font(name="游明朝", size=10)
    font_block_label = Font(name="DF特太ゴシック体", size=20)
    font_label_small = Font(name="游ゴシック", size=11)
    font_subject = Font(name="ＭＳ 明朝", size=12)

    _apply_merge_borders(ws, start_row, border_all, day_periods)

    row1 = start_row
    row2 = start_row + 1
    row3 = start_row + 2
    row32 = start_row + 31
    row33 = start_row + 32
    row34 = start_row + 33
    row49 = start_row + 48
    row50 = start_row + 49

    cell = _safe_cell(ws, row1, 1)
    if cell is not None:
        cell.value = "=TODAY()"
        cell.font = font_date
        cell.alignment = align_center
        cell.number_format = "mm-dd-yy"

    cell = _safe_cell(ws, row1, 4)
    if cell is not None:
        cell.value = title_text
        cell.font = font_title
        cell.alignment = align_center
        cell.border = border_all

    cell = _safe_cell(ws, row2, 1)
    if cell is not None:
        cell.value = scenario_label
        cell.font = font_scenario
        cell.alignment = align_center
        cell.border = border_all

    for day in DAY_ORDER:
        day_range = layout["day_ranges"].get(day)
        if not day_range:
            continue
        cell = _safe_cell(ws, row2, day_range[0])
        if cell is not None:
            cell.value = f"{DAY_JP[day]}曜日"
            cell.font = font_day
            cell.alignment = align_center
            cell.border = border_all

    cell = _safe_cell(ws, row2, 33)
    if cell is not None:
        cell.value = f"=A{row2}"
        cell.font = font_scenario
        cell.alignment = align_center
        cell.border = border_all

    cell = _safe_cell(ws, row33, 1)
    if cell is not None:
        cell.value = f"=A{row2}"
        cell.font = font_block_label
        cell.alignment = align_center
        cell.border = border_all

    for day in DAY_ORDER:
        day_range = layout["day_ranges"].get(day)
        if not day_range:
            continue
        cell = _safe_cell(ws, row33, day_range[0])
        if cell is not None:
            cell.value = f"{DAY_JP[day]}曜日"
            cell.font = font_day
            cell.alignment = align_center
            cell.border = border_all

    cell = _safe_cell(ws, row33, 33)
    if cell is not None:
        cell.value = f"=A{row2}"
        cell.font = font_block_label
        cell.alignment = align_center
        cell.border = border_all

    for day, p, col in layout["slot_columns"]:
        letter = DAY_LETTERS.get(day, day[:1].upper())
        cell = _safe_cell(ws, row32, col)
        if cell is not None:
            cell.value = f"{letter}{p}"
            cell.font = font_day
            cell.alignment = align_center
            cell.border = border_all

    for day, _p, col in layout["slot_columns"]:
            col_letter = get_column_letter(col)
            label_ref = f"{col_letter}{row32}"
            if use_prefix:
                cell = _safe_cell(ws, row3, col)
                if cell is not None:
                    cell.value = f"=$A{row2}&{label_ref}"
                    cell.font = font_time
                    cell.alignment = align_center
                    cell.border = border_all
                cell = _safe_cell(ws, row34, col)
                if cell is not None:
                    cell.value = f"=$A{row2}&{label_ref}"
                    cell.font = font_time
                    cell.alignment = align_center
                    cell.border = border_all
            else:
                cell = _safe_cell(ws, row3, col)
                if cell is not None:
                    cell.value = f"={label_ref}"
                    cell.font = font_time
                    cell.alignment = align_center
                    cell.border = border_all
                cell = _safe_cell(ws, row34, col)
                if cell is not None:
                    cell.value = f"={label_ref}"
                    cell.font = font_time
                    cell.alignment = align_center
                    cell.border = border_all

    for col in range(layout["slot_area_start_col"], RIGHT_LABEL_COL):
        cell = ws.cell(row2, col)
        cell.border = border_all
        if cell.value is None:
            cell.font = font_day
            cell.alignment = align_center
        cell = ws.cell(row3, col)
        cell.border = border_all
        if cell.value is None:
            cell.font = font_time
            cell.alignment = align_center
        cell = ws.cell(row33, col)
        cell.border = border_all
        if cell.value is None:
            cell.font = font_day
            cell.alignment = align_center
        cell = ws.cell(row34, col)
        cell.border = border_all
        if cell.value is None:
            cell.font = font_time
            cell.alignment = align_center

    for r in range(start_row + (TEACHER_ROW_START - 1), start_row + (TEACHER_ROW_END - 1) + 1):
        cell = ws.cell(r, 33)
        _set_value(cell, f"=C{r}")
        cell.font = font_subject
        cell.alignment = align_center
        cell.border = border_all
        cell = ws.cell(r, 34)
        _set_value(cell, f"=B{r}")
        cell.font = font_subject
        cell.alignment = align_center
        cell.border = border_all
    for r in range(start_row + (CLASS_ROW_START - 1), start_row + (CLASS_ROW_END - 1) + 1):
        cell = ws.cell(r, 33)
        _set_value(cell, f"=C{r}")
        cell.font = font_subject
        cell.alignment = align_center
        cell.border = border_all
        cell = ws.cell(r, 34)
        _set_value(cell, f"=A{r}")
        cell.font = font_subject
        cell.alignment = align_center
        cell.border = border_all

    cell = ws.cell(row49, 1)
    _set_value(cell, "特別支援")
    cell.font = font_subject
    cell.alignment = align_center
    cell.border = border_all
    cell = ws.cell(row49, 33)
    _set_value(cell, f"=A{row49}")
    cell.font = font_subject
    cell.alignment = align_center
    cell.border = border_all
    for _day, _p, col in layout["slot_columns"]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row49, col)
        _set_value(cell, f'=IFERROR(INDEX($C${start_row+3}:$AF${start_row+30},MATCH("特",{col_letter}${start_row+3}:{col_letter}${start_row+30},0),1),"")')
        cell.font = font_grid_alt
        cell.alignment = align_center
        cell.border = border_all

    cell = ws.cell(row50, 1)
    _set_value(cell, "巡回(施錠)")
    cell.font = font_subject
    cell.alignment = align_center
    cell.border = border_all
    cell = ws.cell(row50, 33)
    _set_value(cell, f"=A{row50}")
    cell.font = font_subject
    cell.alignment = align_center
    cell.border = border_all
    for _day, _p, col in layout["slot_columns"]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row50, col)
        _set_value(cell, f'=IFERROR(INDEX($C${start_row+3}:$AF${start_row+30},MATCH("巡回",{col_letter}${start_row+3}:{col_letter}${start_row+30},0),1),"")')
        cell.font = font_grid_alt
        cell.alignment = align_center
        cell.border = border_all

    for col in (1, 2, 3):
        cell = _safe_cell(ws, row33, col)
        if cell is not None:
            cell.border = border_all
            cell.alignment = align_center
            cell.font = font_block_label if col == 1 else font_subject
        cell = _safe_cell(ws, row34, col)
        if cell is not None:
            cell.border = border_left_bottom if col == 1 else border_bottom
            cell.alignment = align_center
            cell.font = font_label_small


def _build_yobi_sheet(ws, classes: List[str], day_periods: Dict[str, int]) -> None:
    layout = _build_sheet_layout(day_periods)

    def write_block(start_row: int, title: str) -> None:
        ws.cell(start_row, 1, title)
        for d in DAY_ORDER:
            day_range = layout["day_ranges"].get(d)
            if not day_range:
                continue
            c0 = day_range[0]
            ws.cell(start_row, c0, f"{DAY_JP[d]}曜日")
            for p in range(1, layout["counts"][d] + 1):
                ws.cell(start_row + 1, c0 + p - 1, p)
        for i, c in enumerate(classes):
            ws.cell(start_row + 2 + i, 1, c)

    write_block(1, "音美オンビ")
    write_block(15, "総合ソウゴウ")


def _build_kansei_sheet(ws, classes: List[str], day_periods: Dict[str, int]) -> None:
    ws.cell(1, 1, "完成")
    ws.cell(2, 1, "シナリオ")
    ws.cell(2, 2, "クラス")
    col = 3
    for d in DAY_ORDER:
        for p in range(1, int(day_periods[d]) + 1):
            ws.cell(2, col, f"{DAY_JP[d]}-{p}")
            col += 1


def _build_skill_sheet(ws, classes: List[str], day_periods: Dict[str, int]) -> None:
    ws.cell(1, 1, "技能科目")
    ws.cell(2, 1, "クラス")
    col = 2
    for d in DAY_ORDER:
        for p in range(1, int(day_periods[d]) + 1):
            ws.cell(2, col, f"{DAY_JP[d]}-{p}")
            col += 1
    for i, c in enumerate(classes):
        ws.cell(3 + i, 1, c)


def _find_scenario_id(config: dict, block: str) -> str | None:
    legacy_map = {
        "r6_upper": "1grade-music-art",
        "r6_lower": "1grade-general",
        "１年音美": "1grade-music-art",
        "１年総合": "1grade-general",
    }
    for sc in config.get("scenarios", []):
        if sc.get("target_block") == block:
            return sc.get("id")
    for sc in config.get("scenarios", []):
        sid = str(sc.get("id", "")).strip()
        if sid in legacy_map:
            return sid
        if block == "upper" and sid.endswith("upper"):
            return sid
        if block == "lower" and sid.endswith("lower"):
            return sid
    return None


def _collect_tt_marks(config: dict, scenario_id: str | None) -> set[tuple[str, str]]:
    if not scenario_id:
        return set()
    for sc in config.get("scenarios", []):
        if sc.get("id") == scenario_id:
            marks = sc.get("tt_marks", [])
            return {(m.get("class", ""), m.get("slot", "")) for m in marks if m.get("class") and m.get("slot")}
    return set()


def _teacher_grade_label(class_assignments: list[dict]) -> str:
    grades = set()
    for ca in class_assignments:
        cls = str(ca.get("class", "")).strip()
        hours = ca.get("hours", 0)
        if cls and (hours or ca.get("tt", False)):
            grades.add(grade_of(cls))
    if len(grades) == 1:
        g = next(iter(grades))
        if g.isdigit():
            return f"{g}年"
    return ""


def _build_teacher_rows(config: dict) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    for t in config.get("teachers", []):
        name = str(t.get("name", "")).strip()
        if not name:
            continue
        subjects = t.get("subjects", [])
        subject = str(subjects[0]).strip() if subjects else ""
        grade_label = _teacher_grade_label(t.get("class_assignments", []))
        rows.append((grade_label, subject, name))
    return rows


def _config_classes(config: dict) -> list[str]:
    return [str(cls).strip() for cls in cast(list[Any], config.get("classes", CLASS_ORDER_DEFAULT)) if str(cls).strip()]


def _config_scenarios(config: dict) -> list[dict[str, Any]]:
    return [cast(dict[str, Any], scenario) for scenario in cast(list[Any], config.get("scenarios", [])) if isinstance(scenario, dict)]


def _config_day_periods(config: dict) -> dict[str, int]:
    raw = cast(dict[str, Any], config.get("day_periods", {}))
    return {day: int(raw.get(day, 0) or 0) for day in DAY_ORDER}


def _build_teacher_slot_marks(config: dict) -> dict[str, dict[str, str]]:
    teacher_slot_marks: dict[str, dict[str, str]] = {}
    for teacher in config.get("teachers", []):
        name = str(teacher.get("name", "")).strip()
        if not name:
            continue
        slot_marks = teacher_slot_marks.setdefault(name, {})
        for slot in teacher.get("discouraged_slots", []):
            slot_key = str(slot).strip()
            if slot_key:
                slot_marks[slot_key] = EXCEL_UNAVAIL_SOFT_MARK
        for slot in teacher.get("unavailable_slots", []):
            slot_key = str(slot).strip()
            if slot_key:
                slot_marks[slot_key] = EXCEL_UNAVAIL_HARD_MARK
        if not slot_marks:
            teacher_slot_marks.pop(name, None)
    return teacher_slot_marks


def _build_individual_teacher_label(config: dict) -> str:
    names: list[str] = []
    for teacher in config.get("teachers", []):
        name = str(teacher.get("name", "")).strip()
        if not name:
            continue
        subjects = [
            _normalize_subject_name(str(subject))
            for subject in teacher.get("subjects", [])
            if _normalize_subject_name(str(subject))
        ]
        if "個別" in subjects and name not in names:
            names.append(name)
    return "・".join(names)


def _ensure_class_subject_teacher(config: dict) -> dict:
    existing: dict[str, dict[str, str]] = {}
    raw_existing = config.get("class_subject_teacher", {})
    if isinstance(raw_existing, dict):
        for cls, subj_map in raw_existing.items():
            if not isinstance(subj_map, dict):
                continue
            existing[str(cls).strip()] = {
                str(subj).strip(): str(name).strip()
                for subj, name in subj_map.items()
                if str(subj).strip() and str(name).strip()
            }
    merged: dict[str, dict[str, str]] = {}
    for t in config.get("teachers", []):
        name = str(t.get("name", "")).strip()
        if not name:
            continue
        subjects = [str(s).strip() for s in t.get("subjects", []) if str(s).strip()]
        if not subjects:
            continue
        for ca in t.get("class_assignments", []):
            cls = str(ca.get("class", "")).strip()
            hours = ca.get("hours", 0)
            if not cls or (hours is None and not ca.get("tt", False)):
                continue
            if hours == 0 and not ca.get("tt", False):
                continue
            for subj in subjects:
                merged.setdefault(cls, {})[subj] = name
    for cls, subj_map in existing.items():
        for subj, name in subj_map.items():
            merged.setdefault(cls, {}).setdefault(subj, name)
    return merged


def _build_class_subject_teacher_index(config: dict) -> dict:
    index: dict[str, dict[str, dict[str, list[str]]]] = {"__all__": {}}
    for t in config.get("teachers", []):
        name = str(t.get("name", "")).strip()
        if not name:
            continue
        subjects = [str(s).strip() for s in t.get("subjects", []) if str(s).strip()]
        if not subjects:
            continue
        for subj in subjects:
            entry = index["__all__"].setdefault(subj, {"tt": [], "regular": []})
            if name not in entry["regular"]:
                entry["regular"].append(name)
        for ca in t.get("class_assignments", []):
            cls = str(ca.get("class", "")).strip()
            hours = ca.get("hours", 0)
            if not cls or (hours is None and not ca.get("tt", False)):
                continue
            if hours == 0 and not ca.get("tt", False):
                continue
            bucket = "tt" if ca.get("tt", False) else "regular"
            for subj in subjects:
                entry = index.setdefault(cls, {}).setdefault(subj, {"tt": [], "regular": []})
                if name not in entry[bucket]:
                    entry[bucket].append(name)
    return index


def _build_tt_assignment_index(
    config: dict,
    assignments: Dict[str, Dict[str, str]],
    tt_marks: set[tuple[str, str]] | None,
    class_subject_teacher: Dict[str, Dict[str, str]],
    class_subject_teacher_index: dict | None = None,
) -> dict[tuple[str, str], list[str]]:
    marks = {
        (str(cls).strip(), str(slot).strip())
        for cls, slot in (tt_marks or set())
        if str(cls).strip() and str(slot).strip()
    }
    if not marks:
        return {}

    def _slot_sort_key(slot: str) -> tuple[int, int, str]:
        try:
            day, period = parse_slot_key(slot).day, parse_slot_key(slot).period
            day_idx = DAY_ORDER.index(day) if day in DAY_ORDER else len(DAY_ORDER)
            return day_idx, int(period), slot
        except Exception:
            return len(DAY_ORDER), 999, slot

    def _resolve_regular_teachers(cls: str, subj: str) -> list[str]:
        regular_teachers: list[str] = []
        if class_subject_teacher_index is not None:
            entry = class_subject_teacher_index.get(cls, {}).get(subj)
            if entry:
                regular_teachers = list(entry.get("regular", []))
        if not regular_teachers:
            fallback = str(class_subject_teacher.get(cls, {}).get(subj, "")).strip()
            if fallback:
                regular_teachers = [fallback]
        if class_subject_teacher_index is not None and not regular_teachers:
            global_entry = class_subject_teacher_index.get("__all__", {}).get(subj)
            if global_entry:
                global_regular = list(global_entry.get("regular", []))
                if len(global_regular) == 1:
                    regular_teachers = global_regular
        return regular_teachers

    teacher_busy_slots: dict[str, set[str]] = {}
    for cls, slots in assignments.items():
        for slot, subject in slots.items():
            subj = _normalize_subject_name(str(subject).replace("(T)", ""))
            if not subj:
                continue
            lookup_subjects = [subj]
            if subj == "音美":
                lookup_subjects = ["音楽", "美術"]
            elif subj == "技家":
                lookup_subjects = ["技術", "家庭"]
            for lookup_subj in lookup_subjects:
                for teacher in _resolve_regular_teachers(cls, lookup_subj):
                    teacher_busy_slots.setdefault(teacher, set()).add(slot)

    tt_assignment_index: dict[tuple[str, str], list[str]] = {}

    for teacher_data in config.get("teachers", []):
        teacher = str(teacher_data.get("name", "")).strip()
        if not teacher:
            continue
        subjects = [str(s).strip() for s in teacher_data.get("subjects", []) if str(s).strip()]
        subject = _normalize_subject_name(subjects[0]) if subjects else ""
        if not subject:
            continue
        unavailable_slots = {str(slot).strip() for slot in teacher_data.get("unavailable_slots", []) if str(slot).strip()}

        units: list[tuple[str, list[str]]] = []
        for class_assignment in teacher_data.get("class_assignments", []):
            cls = str(class_assignment.get("class", "")).strip()
            if not cls or not bool(class_assignment.get("tt", False)):
                continue
            try:
                hours = int(class_assignment.get("hours", 0))
            except Exception:
                try:
                    hours = int(float(class_assignment.get("hours", 0)))
                except Exception:
                    hours = 0
            if hours <= 0:
                continue

            candidates: list[str] = []
            for slot, assigned_subject in assignments.get(cls, {}).items():
                key = (cls, slot)
                if key not in marks:
                    continue
                normalized_assigned = _normalize_subject_name(str(assigned_subject).replace("(T)", ""))
                if normalized_assigned != subject:
                    continue
                if slot in unavailable_slots:
                    continue
                if slot in teacher_busy_slots.get(teacher, set()):
                    continue
                candidates.append(slot)

            candidates.sort(key=_slot_sort_key)
            for _ in range(hours):
                units.append((cls, candidates))

        if not units:
            continue

        def _unit_order_key(unit_idx: int) -> tuple[int, tuple, int]:
            cls_name, candidates = units[unit_idx]
            return len(candidates), _class_sort_key(cls_name), unit_idx

        unit_order = sorted(range(len(units)), key=_unit_order_key)
        matched_slot_to_unit: dict[str, int] = {}

        def _try_match(unit_idx: int, seen_slots: set[str]) -> bool:
            cls, candidates = units[unit_idx]
            for slot in candidates:
                if slot in seen_slots:
                    continue
                seen_slots.add(slot)
                prev = matched_slot_to_unit.get(slot)
                if prev is None or _try_match(prev, seen_slots):
                    matched_slot_to_unit[slot] = unit_idx
                    return True
            return False

        for unit_idx in unit_order:
            _try_match(unit_idx, set())

        for slot, unit_idx in matched_slot_to_unit.items():
            cls, _candidates = units[unit_idx]
            tt_assignment_index.setdefault((cls, slot), []).append(teacher)

    for key in tt_assignment_index:
        tt_assignment_index[key] = sorted(dict.fromkeys(tt_assignment_index[key]))
    return tt_assignment_index


def _build_teacher_schedule(
    assignments: Dict[str, Dict[str, str]],
    origin_subjects: Dict[str, Dict[str, str]],
    class_subject_teacher: Dict[str, Dict[str, str]],
    tt_marks: set[tuple[str, str]] | None = None,
    class_subject_teacher_index: dict | None = None,
    tt_assignment_index: dict[tuple[str, str], list[str]] | None = None,
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    teacher_schedule: dict[str, dict[str, str]] = {}
    skip_subjects = {"道徳", "学活", "個別"}
    marks = tt_marks or set()
    teacher_slot_classes: dict[str, dict[str, list[str]]] = {}
    teacher_slot_origins: dict[str, dict[str, set[str]]] = {}

    def _add(teacher: str, slot: str, cls: str, origin_subject: str | None = None) -> None:
        if not teacher:
            return
        slots = teacher_slot_classes.setdefault(teacher, {})
        cls_list = slots.setdefault(slot, [])
        if cls not in cls_list:
            cls_list.append(cls)
        normalized_origin = _normalize_subject_name(origin_subject or "")
        if normalized_origin in {"音美", "技家"}:
            teacher_slot_origins.setdefault(teacher, {}).setdefault(slot, set()).add(normalized_origin)

    def _format_classes(classes: list[str]) -> str:
        uniq = list(dict.fromkeys(classes))
        uniq.sort(key=_class_sort_key)
        grades = []
        nums = []
        for c in uniq:
            parts = c.split("-")
            if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
                return ",".join(uniq)
            grades.append(parts[0])
            nums.append(int(parts[1]))
        if len(set(grades)) != 1:
            return ",".join(uniq)
        nums.sort()
        head = f"{grades[0]}-{nums[0]}"
        if len(nums) == 1:
            return head
        tail = ",".join(str(n) for n in nums[1:])
        return f"{head},{tail}"

    def _resolve_teachers(cls: str, subj: str) -> tuple[list[str], list[str]]:
        regular_teachers: list[str] = []
        tt_teachers: list[str] = []
        if class_subject_teacher_index is not None:
            entry = class_subject_teacher_index.get(cls, {}).get(subj)
            if entry:
                regular_teachers = list(entry.get("regular", []))
                tt_teachers = list(entry.get("tt", []))
        if not regular_teachers:
            fallback = class_subject_teacher.get(cls, {}).get(subj, "").strip()
            if fallback:
                regular_teachers = [fallback]
        if class_subject_teacher_index is not None:
            global_entry = class_subject_teacher_index.get("__all__", {}).get(subj)
            if global_entry:
                if not regular_teachers:
                    global_regular = list(global_entry.get("regular", []))
                    if len(global_regular) == 1:
                        regular_teachers = global_regular
                if not tt_teachers:
                    global_tt = list(global_entry.get("tt", []))
                    if len(global_tt) == 1:
                        tt_teachers = global_tt
        return regular_teachers, tt_teachers

    teacher_schedule_origins: dict[str, dict[str, str]] = {}

    for cls, slots in assignments.items():
        for slot, subject in slots.items():
            subj = str(subject).strip()
            if not subj or subj in skip_subjects:
                continue
            origin_subject = origin_subjects.get(cls, {}).get(slot)
            lookup_subjects = [subj]
            if subj == "音美":
                lookup_subjects = ["音楽", "美術"]
                origin_subject = "音美"
            elif subj == "技家":
                lookup_subjects = ["技術", "家庭"]
                origin_subject = "技家"

            for lookup_subj in lookup_subjects:
                regular_teachers, tt_teachers = _resolve_teachers(cls, lookup_subj)
                for t in regular_teachers:
                    _add(t, slot, cls, origin_subject)
                if tt_assignment_index is not None:
                    for t in tt_assignment_index.get((cls, slot), []):
                        _add(t, slot, cls, origin_subject)
                elif (cls, slot) in marks:
                    for t in tt_teachers:
                        _add(t, slot, cls, origin_subject)

    for teacher, slots in teacher_slot_classes.items():
        for slot, cls_list in slots.items():
            if cls_list:
                teacher_schedule.setdefault(teacher, {})[slot] = _format_classes(cls_list)
                origin_set = teacher_slot_origins.get(teacher, {}).get(slot, set())
                if len(origin_set) == 1:
                    teacher_schedule_origins.setdefault(teacher, {})[slot] = next(iter(origin_set))
    return teacher_schedule, teacher_schedule_origins


def _build_variant_assignments(
    upper: Dict[str, Dict[str, str]],
    lower: Dict[str, Dict[str, str]],
    classes: list[str],
    use_lower: bool,
    onbi_start_music: bool | None,
    tech_subject: str | None,
    merge_general_from_lower: bool,
    onbi_overrides: dict[str, str] | None = None,
    tech_overrides: dict[str, str] | None = None,
) -> tuple[Dict[str, Dict[str, str]], Dict[str, Dict[str, str]]]:
    base = lower if use_lower else upper
    out: Dict[str, Dict[str, str]] = {c: {} for c in classes}
    origins: Dict[str, Dict[str, str]] = {c: {} for c in classes}
    grade1_classes = [c for c in classes if str(c).startswith("1-")]
    grade1_classes.sort(key=_class_sort_key)
    onbi_map: dict[str, str] = {}
    if onbi_start_music is not None:
        for idx, cls in enumerate(grade1_classes):
            if onbi_start_music:
                onbi_map[cls] = "音楽" if idx % 2 == 0 else "美術"
            else:
                onbi_map[cls] = "美術" if idx % 2 == 0 else "音楽"
    onbi_overrides = onbi_overrides or {}
    tech_overrides = tech_overrides or {}

    for cls, slots in base.items():
        for slot, subject in slots.items():
            subj = str(subject).strip()
            if not subj:
                continue
            origin_subject = _normalize_subject_name(subj)
            if merge_general_from_lower and subj == "総合":
                alt = str(lower.get(cls, {}).get(slot, "")).strip()
                if alt and alt != "総合":
                    subj = alt
            if subj == "音美":
                override = onbi_overrides.get(f"{cls}|{slot}")
                if override:
                    subj = override
                elif cls in onbi_map:
                    subj = onbi_map[cls]
            elif subj == "技家":
                override = tech_overrides.get(f"{cls}|{slot}")
                if override:
                    subj = override
                elif tech_subject:
                    subj = tech_subject
            out.setdefault(cls, {})[slot] = subj
            if origin_subject in {"音美", "技家"}:
                origins.setdefault(cls, {})[slot] = origin_subject
    return out, origins


def _default_excel_variants() -> list[dict]:
    return [
        {"id": "①", "use_lower": False, "onbi_start_music": True, "tech_subject": "技術", "onbi_overrides": {}, "tech_overrides": {}},
        {"id": "②", "use_lower": False, "onbi_start_music": False, "tech_subject": "家庭", "onbi_overrides": {}, "tech_overrides": {}},
        {"id": "③", "use_lower": True, "onbi_start_music": None, "tech_subject": "技術", "onbi_overrides": {}, "tech_overrides": {}},
        {"id": "④", "use_lower": False, "onbi_start_music": True, "tech_subject": "家庭", "onbi_overrides": {}, "tech_overrides": {}},
        {"id": "⑤", "use_lower": False, "onbi_start_music": False, "tech_subject": "技術", "onbi_overrides": {}, "tech_overrides": {}},
        {"id": "⑥", "use_lower": True, "onbi_start_music": None, "tech_subject": "家庭", "onbi_overrides": {}, "tech_overrides": {}},
    ]


def _apply_excel_variant_overrides(config: dict, variants: list[dict]) -> list[dict]:
    overrides = config.get("excel_variant_settings", {}) or {}
    for v in variants:
        key = v.get("id")
        cfg = overrides.get(key, {})
        if "onbi_start_music" in cfg:
            v["onbi_start_music"] = cfg["onbi_start_music"]
        if "tech_subject" in cfg:
            v["tech_subject"] = cfg["tech_subject"]
        if "onbi_overrides" in cfg:
            v["onbi_overrides"] = dict(cfg["onbi_overrides"])
        if "tech_overrides" in cfg:
            v["tech_overrides"] = dict(cfg["tech_overrides"])
    return variants


def _variant_title(variant_id: str, use_lower: bool, onbi_start_music: bool | None, tech_subject: str | None) -> str:
    if use_lower:
        first = "(1年)総"
    else:
        first = f"(1年{'音' if onbi_start_music else '美'})"
    third = f"(3年{'技' if tech_subject == '技術' else '家'})"
    return f"＜先生の授業時間割一覧　{variant_id}{first}{third}＞"


def _fill_block(
    ws,
    start_row: int,
    assignments: Dict[str, Dict[str, str]],
    origin_subjects: Dict[str, Dict[str, str]],
    tt_marks: set[tuple[str, str]],
    classes: list[str],
    teacher_rows: list[tuple[str, str, str]],
    class_subject_teacher: Dict[str, Dict[str, str]],
    class_subject_teacher_index: dict,
    day_periods: Dict[str, int],
    config: dict,
) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    slot_columns = _iter_slot_columns(day_periods)
    slot_to_col_map = {f"{d}-{p}": col for d, p, col in slot_columns}
    individual_teacher_label = _build_individual_teacher_label(config)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    font_subject = Font(name="ＭＳ 明朝", size=12)
    font_grid = Font(name="游ゴシック", size=10)

    teacher_row_start = start_row + (TEACHER_ROW_START - 1)
    teacher_row_end = start_row + (TEACHER_ROW_END - 1)
    class_row_start = start_row + (CLASS_ROW_START - 1)
    class_row_end = start_row + (CLASS_ROW_END - 1)

    clear_fill = PatternFill()
    pe_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    onbi_fill = PatternFill(start_color=ONBI_FILL_COLOR, end_color=ONBI_FILL_COLOR, fill_type="solid")
    techka_fill = PatternFill(start_color=TECHKA_FILL_COLOR, end_color=TECHKA_FILL_COLOR, fill_type="solid")
    unavailable_fill = PatternFill(start_color=EXCEL_UNAVAIL_HARD_FILL_COLOR, end_color=EXCEL_UNAVAIL_HARD_FILL_COLOR, fill_type="solid")
    discouraged_fill = PatternFill(start_color=EXCEL_UNAVAIL_SOFT_FILL_COLOR, end_color=EXCEL_UNAVAIL_SOFT_FILL_COLOR, fill_type="solid")
    teacher_slot_marks = _build_teacher_slot_marks(config)

    def _group_fill(group_key: str | None):
        if group_key == "onbi":
            return onbi_fill
        if group_key == "techka":
            return techka_fill
        return None

    def _teacher_slot_fill(teacher_name: str, slot: str):
        mark = teacher_slot_marks.get(teacher_name, {}).get(slot)
        if mark == EXCEL_UNAVAIL_HARD_MARK:
            return unavailable_fill
        if mark == EXCEL_UNAVAIL_SOFT_MARK:
            return discouraged_fill
        return None

    for r in range(teacher_row_start, teacher_row_end + 1):
        for _day, _p, col in slot_columns:
            cell = ws.cell(r, col)
            cell.value = None
            cell.fill = clear_fill
            cell.border = border_all
            cell.font = font_grid
            cell.alignment = align_center

    for idx in range(teacher_row_start, teacher_row_end + 1):
        for col in (1, 2, 3):
            cell = ws.cell(idx, col)
            _set_value(cell, None)
            cell.border = border_all
            cell.font = font_subject
            cell.alignment = align_center
    for i, (grade_label, subject, name) in enumerate(teacher_rows[: (teacher_row_end - teacher_row_start + 1)]):
        r = teacher_row_start + i
        cell = ws.cell(r, 1)
        _set_value(cell, grade_label)
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center
        cell = ws.cell(r, 2)
        _set_value(cell, subject)
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center
        cell = ws.cell(r, 3)
        _set_value(cell, name)
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center
        for slot, col in slot_to_col_map.items():
            mark_fill = _teacher_slot_fill(name, slot)
            if mark_fill is not None:
                ws.cell(r, col).fill = mark_fill

    tt_assignment_index = _build_tt_assignment_index(
        config=config,
        assignments=assignments,
        tt_marks=tt_marks,
        class_subject_teacher=class_subject_teacher,
        class_subject_teacher_index=class_subject_teacher_index,
    )

    teacher_schedule, teacher_schedule_origins = _build_teacher_schedule(
        assignments,
        origin_subjects,
        class_subject_teacher,
        tt_marks,
        class_subject_teacher_index,
        tt_assignment_index,
    )
    teacher_index = {name: teacher_row_start + i for i, (_g, _s, name) in enumerate(teacher_rows)}

    for teacher, slots in teacher_schedule.items():
        r = teacher_index.get(teacher)
        if r is None:
            continue
        for slot, cls in slots.items():
            col = slot_to_col_map.get(slot)
            if not col:
                continue
            cell = ws.cell(r, col)
            cell.value = cls
            cell.number_format = "@"
            cell.border = border_all
            cell.font = font_grid
            cell.alignment = align_center
            mark_fill = _teacher_slot_fill(teacher, slot)
            teacher_fill = _group_fill(_origin_group_key(teacher_schedule_origins.get(teacher, {}).get(slot)))
            if mark_fill is not None:
                cell.fill = mark_fill
            elif teacher_fill is not None:
                cell.fill = teacher_fill

    for r in range(class_row_start, class_row_end + 1):
        for col in (1, 2, 3):
            cell = ws.cell(r, col)
            _set_value(cell, None)
            cell.border = border_all
            cell.font = font_subject
            cell.alignment = align_center
        for _day, _p, col in slot_columns:
            cell = ws.cell(r, col)
            cell.value = None
            cell.fill = clear_fill
            cell.border = border_all
            cell.font = font_grid
            cell.alignment = align_center

    for i, cls in enumerate(classes):
        r = class_row_start + i
        if r > class_row_end:
            break
        cell = ws.cell(r, 1)
        _set_value(cell, cls)
        cell.number_format = "@"
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center
        cell = ws.cell(r, 3)
        _set_value(cell, "")
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center

        for slot, subject in assignments.get(cls, {}).items():
            col = slot_to_col_map.get(slot)
            if not col:
                continue
            subj = str(subject).strip()
            if not subj:
                continue
            if subj.endswith("(T)"):
                subj = subj[:-3]
            if _normalize_subject_name(subj) == "個別":
                continue
            cell = ws.cell(r, col)
            origin_fill = _group_fill(_origin_group_key(origin_subjects.get(cls, {}).get(slot)))
            if origin_fill is not None:
                cell.value = subj
                cell.fill = origin_fill
            elif subj == "保体":
                cell.value = subj
                cell.fill = pe_fill
            else:
                cell.value = subj
            cell.border = border_all
            cell.font = font_grid
            cell.alignment = align_center

    individual_row = class_row_start + len(classes)
    if individual_row <= class_row_end:
        cell = ws.cell(individual_row, 1)
        _set_value(cell, "個別")
        cell.number_format = "@"
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center
        cell = ws.cell(individual_row, 3)
        _set_value(cell, individual_teacher_label)
        cell.number_format = "@"
        cell.border = border_all
        cell.font = font_subject
        cell.alignment = align_center


def _create_workbook_simple(config: dict, solved_by_scenario: Dict[str, Dict[str, Dict[str, str]]], output_path: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。") from e

    classes = _config_classes(config)
    day_periods = _config_day_periods(config)
    scenarios = _config_scenarios(config)

    wb = cast(Any, openpyxl.Workbook())
    ws0 = cast(Any, wb.active)
    ws0.title = "時間割（略図）"
    wb.create_sheet("完成")
    wb.create_sheet("技能科目")
    wb.create_sheet("予備")

    ws_kansei = cast(Any, wb["完成"])
    ws_skill = cast(Any, wb["技能科目"])
    ws_yobi = cast(Any, wb["予備"])

    _build_kansei_sheet(ws_kansei, classes, day_periods)
    _build_skill_sheet(ws_skill, classes, day_periods)
    _build_yobi_sheet(ws_yobi, classes, day_periods)

    for scenario in scenarios:
        sid = str(scenario.get("id", "")).strip()
        block = str(scenario.get("target_block", "upper") or "upper")
        row_base = row_base_for_block(block)
        solved = solved_by_scenario[sid]
        for idx, c in enumerate(classes):
            row = row_base + idx
            for slot_key, subj in solved.get(c, {}).items():
                if _normalize_subject_name(subj) == "個別":
                    continue
                sl = parse_slot_key(slot_key)
                col = slot_to_col(sl, day_periods)
                ws_yobi.cell(row, col, subj)

    row = 3
    for scenario in scenarios:
        sid = str(scenario.get("id", "")).strip()
        solved = solved_by_scenario[sid]
        for c in classes:
            ws_kansei.cell(row, 1, sid)
            ws_kansei.cell(row, 2, c)
            col = 3
            for d in DAY_ORDER:
                for p in range(1, int(day_periods[d]) + 1):
                    subj = solved.get(c, {}).get(f"{d}-{p}", "")
                    ws_kansei.cell(row, col, "" if _normalize_subject_name(subj) == "個別" else subj)
                    col += 1
            row += 1

    skill_subjects = set(config.get("skill_subjects", []))
    if scenarios:
        first_id = str(scenarios[0].get("id", "")).strip()
        first = solved_by_scenario[first_id]
        for i, c in enumerate(classes):
            row_skill = 3 + i
            col = 2
            for d in DAY_ORDER:
                for p in range(1, int(day_periods[d]) + 1):
                    subj = first.get(c, {}).get(f"{d}-{p}", "")
                    ws_skill.cell(row_skill, col, subj if subj in skill_subjects else "")
                    col += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def create_workbook_by_structure(config: dict, solved_by_scenario: Dict[str, Dict[str, Dict[str, str]]], output_path: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。") from e

    if config.get("year") != 6:
        _create_workbook_simple(config, solved_by_scenario, output_path)
        return

    classes = _config_classes(config)
    day_periods = _config_day_periods(config)
    teacher_rows = _build_teacher_rows(config)
    class_subject_teacher = _ensure_class_subject_teacher(config)
    class_subject_teacher_index = _build_class_subject_teacher_index(config)

    teacher_name_max = max((len(name) for _g, _s, name in teacher_rows), default=6)
    col_c_width = max(8.0, min(16.0, teacher_name_max + 2))
    slot_count = _build_sheet_layout(day_periods)["total_slots"]

    page_width_in = 16.54
    page_height_in = 11.69
    margin_in = 0.25
    printable_width_in = page_width_in - (margin_in * 2)
    printable_height_in = page_height_in - (margin_in * 2)

    printable_width_px = printable_width_in * 96.0
    fixed_cols = [5.5, 6.0, col_c_width, 1.0, 1.0, 6.0, 6.0]
    fixed_width_px = sum(w * 7.0 + 5.0 for w in fixed_cols)
    slot_width_px = (printable_width_px - fixed_width_px) / max(1, slot_count)
    slot_width = max(3.0, (slot_width_px - 5.0) / 7.0)

    total_grid_rows = (TEACHER_ROW_END - TEACHER_ROW_START + 1) + (CLASS_ROW_END - CLASS_ROW_START + 1) + 2
    printable_height_pt = printable_height_in * 72.0
    fixed_height = 15.0 + 12.75 + 12.75 + 0.75 + 12.0 + 12.0 + 4.5
    grid_row_height = max(12.0, (printable_height_pt - fixed_height) / max(1, total_grid_rows))

    upper_id = _find_scenario_id(config, "upper")
    lower_id = _find_scenario_id(config, "lower")
    upper = solved_by_scenario.get(upper_id or "", {})
    lower = solved_by_scenario.get(lower_id or "", {})

    tt_upper = _collect_tt_marks(config, upper_id)
    tt_lower = _collect_tt_marks(config, lower_id)

    wb = cast(Any, openpyxl.Workbook())
    active = wb.active
    if active is not None:
        wb.remove(active)

    variants = cast(list[dict[str, Any]], _apply_excel_variant_overrides(config, _default_excel_variants()))

    ws_kansei = cast(Any, wb.create_sheet("完成"))
    _apply_dimensions(ws_kansei, slot_width=slot_width, col_c_width=col_c_width, day_periods=day_periods)

    print_areas: list[str] = []
    for idx, v in enumerate(variants):
        name = v["id"]
        use_lower = bool(v.get("use_lower"))
        onbi_start_music = v.get("onbi_start_music")
        tech_subject = v.get("tech_subject")
        onbi_overrides = v.get("onbi_overrides") or {}
        tech_overrides = v.get("tech_overrides") or {}
        title = _variant_title(name, use_lower, onbi_start_music, tech_subject)
        start_row = 1 + (idx * BLOCK_PITCH)
        _init_block(ws_kansei, start_row, name, title, day_periods, use_prefix=True, grid_row_height=grid_row_height)
        print_areas.append(f"$A${start_row}:$AH${start_row + (BLOCK_HEIGHT - 1)}")

        assignments, origin_subjects = _build_variant_assignments(
            upper=upper,
            lower=lower,
            classes=classes,
            use_lower=use_lower,
            onbi_start_music=onbi_start_music,
            tech_subject=tech_subject,
            merge_general_from_lower=not use_lower,
            onbi_overrides=onbi_overrides,
            tech_overrides=tech_overrides,
        )
        tt_marks = tt_lower if use_lower else tt_upper
        _fill_block(
            ws_kansei,
            start_row=start_row,
            assignments=assignments,
            origin_subjects=origin_subjects,
            tt_marks=tt_marks,
            classes=classes,
            teacher_rows=teacher_rows,
            class_subject_teacher=class_subject_teacher,
            class_subject_teacher_index=class_subject_teacher_index,
            day_periods=day_periods,
            config=config,
        )

    ws_kansei.print_area = ",".join(print_areas)
    ws_kansei.sheet_view.view = "pageBreakPreview"
    ws_kansei.page_setup.orientation = ws_kansei.ORIENTATION_LANDSCAPE
    ws_kansei.page_setup.paperSize = ws_kansei.PAPERSIZE_A3
    ws_kansei.page_setup.fitToWidth = 1
    ws_kansei.page_setup.fitToHeight = 6
    ws_kansei.page_setup.fitToPage = True
    ws_kansei.page_margins.left = margin_in
    ws_kansei.page_margins.right = margin_in
    ws_kansei.page_margins.top = margin_in
    ws_kansei.page_margins.bottom = margin_in
    ws_kansei.page_margins.header = 0.1
    ws_kansei.page_margins.footer = 0.1
    try:
        from openpyxl.worksheet.pagebreak import Break

        for idx in range(1, len(variants)):
            ws_kansei.row_breaks.append(Break(id=1 + (idx * BLOCK_PITCH)))
    except Exception:
        pass

    ws_r6 = cast(Any, wb.create_sheet("略図"))
    _apply_dimensions(ws_r6, slot_width=slot_width, col_c_width=col_c_width, day_periods=day_periods)
    _init_block(ws_r6, 1, "①②④⑤", "＜先生の授業時間割一覧　①②④⑤＞", day_periods, use_prefix=False, grid_row_height=grid_row_height)
    _init_block(ws_r6, 1 + BLOCK_PITCH, "③⑥", "＜先生の授業時間割一覧　③⑥＞", day_periods, use_prefix=False, grid_row_height=grid_row_height)
    ws_r6.print_area = f"$A$1:$AH${BLOCK_HEIGHT},$A${1 + BLOCK_PITCH}:$AH${BLOCK_PITCH + BLOCK_HEIGHT}"
    ws_r6.sheet_view.view = "pageBreakPreview"
    ws_r6.page_setup.orientation = ws_r6.ORIENTATION_LANDSCAPE
    ws_r6.page_setup.paperSize = ws_r6.PAPERSIZE_A3
    ws_r6.page_setup.fitToWidth = 1
    ws_r6.page_setup.fitToHeight = 2
    ws_r6.page_setup.fitToPage = True
    ws_r6.page_margins.left = margin_in
    ws_r6.page_margins.right = margin_in
    ws_r6.page_margins.top = margin_in
    ws_r6.page_margins.bottom = margin_in
    ws_r6.page_margins.header = 0.1
    ws_r6.page_margins.footer = 0.1
    try:
        from openpyxl.worksheet.pagebreak import Break
        ws_r6.row_breaks.append(Break(id=1 + BLOCK_PITCH))
    except Exception:
        pass

    ryaku_upper, ryaku_upper_origins = _build_variant_assignments(
        upper=upper,
        lower=lower,
        classes=classes,
        use_lower=False,
        onbi_start_music=None,
        tech_subject=None,
        merge_general_from_lower=True,
        onbi_overrides={},
        tech_overrides={},
    )
    ryaku_lower, ryaku_lower_origins = _build_variant_assignments(
        upper=upper,
        lower=lower,
        classes=classes,
        use_lower=True,
        onbi_start_music=None,
        tech_subject=None,
        merge_general_from_lower=False,
        onbi_overrides={},
        tech_overrides={},
    )
    _fill_block(
        ws_r6,
        start_row=1,
        assignments=ryaku_upper,
        origin_subjects=ryaku_upper_origins,
        tt_marks=tt_upper,
        classes=classes,
        teacher_rows=teacher_rows,
        class_subject_teacher=class_subject_teacher,
        class_subject_teacher_index=class_subject_teacher_index,
        day_periods=day_periods,
        config=config,
    )
    _fill_block(
        ws_r6,
        start_row=1 + BLOCK_PITCH,
        assignments=ryaku_lower,
        origin_subjects=ryaku_lower_origins,
        tt_marks=tt_lower,
        classes=classes,
        teacher_rows=teacher_rows,
        class_subject_teacher=class_subject_teacher,
        class_subject_teacher_index=class_subject_teacher_index,
        day_periods=day_periods,
        config=config,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

